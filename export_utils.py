"""Export DataFrames to Excel or CSV for download."""
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def _build_inner_outer_paired_df(inner_df, full_df, same_entity):
    """Build a DataFrame with INNER/OUTER rows paired by Match_Group. Optimized for large dfs: one groupby pass, one concat."""
    if inner_df is None or inner_df.empty:
        return pd.DataFrame()
    full_df = full_df.copy()
    full_df["_outer_key"] = full_df["gtin_outer_normalized"].astype(str).str.strip()
    data_cols = [c for c in inner_df.columns if c in full_df.columns]
    # Build key -> list of indices in one pass (no get_group per key)
    if same_entity:
        key_to_indices = {}
        for k, g in full_df.groupby(["_outer_key", "Legal Entity"], dropna=False):
            key_to_indices[k] = g.index.tolist()
    else:
        key_to_indices = {}
        for k, g in full_df.groupby("_outer_key", dropna=False):
            key_to_indices[k] = g.index.tolist()
    get_indices = key_to_indices.get
    inner_keys = inner_df["gtin_inner_normalized"].astype(str).str.strip().values
    entities = inner_df["Legal Entity"].values
    n = len(inner_df)
    # Collect all outer indices and match_ids for one big loc
    all_outer_indices = []
    all_outer_match_ids = []
    for match_id in range(n):
        key = inner_keys[match_id]
        row_entity = entities[match_id]
        lookup = (key, row_entity) if same_entity and row_entity is not None else key
        indices = get_indices(lookup, [])
        all_outer_indices.extend(indices)
        all_outer_match_ids.extend([match_id + 1] * len(indices))
    full_df.drop(columns=["_outer_key"], inplace=True, errors="ignore")
    # Build inner block (one DataFrame, no row loop)
    inner_block = inner_df[data_cols].copy()
    inner_block.insert(0, "Match_Group", range(1, n + 1))
    inner_block.insert(0, "Role", "INNER")
    if not all_outer_indices:
        out = inner_block
    else:
        outer_block = full_df.loc[all_outer_indices, data_cols].copy()
        outer_block.insert(0, "Match_Group", all_outer_match_ids)
        outer_block.insert(0, "Role", "OUTER")
        # Interleave INNER/OUTER by Match_Group: sort so each INNER is followed by its OUTER rows
        out = pd.concat([inner_block, outer_block], ignore_index=True)
        out = out.sort_values(by=["Match_Group", "Role"], ascending=[True, True]).reset_index(drop=True)
    col_order = ["Role", "Match_Group"] + data_cols
    return out[[c for c in col_order if c in out.columns]]


0# Below this row count: apply green fill. Above: skip (openpyxl cell-by-cell is O(rows*cols), can take hours).
_INNER_OUTER_STYLE_ROW_LIMIT = 8000


def to_excel_bytes_inner_outer_paired(inner_df, full_df, same_entity=True) -> bytes:
    """Export Inner rows with matching Outer rows, paired by Match_Group, with alternating green fill (skipped for large sheets)."""
    df = _build_inner_outer_paired_df(inner_df, full_df, same_entity)
    if df is None or df.empty:
        buf = io.BytesIO()
        pd.DataFrame(columns=["Role", "Match_Group"]).to_excel(buf, index=False, engine="openpyxl")
        buf.seek(0)
        return buf.getvalue()
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    if len(df) > _INNER_OUTER_STYLE_ROW_LIMIT:
        return buf.getvalue()
    wb = load_workbook(buf)
    ws = wb.active
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    mg = pd.to_numeric(df["Match_Group"], errors="coerce").fillna(0).astype(int)
    odd_rows = {i + 2 for i in range(len(df)) if mg.iloc[i] % 2 == 1}
    for row_idx in odd_rows:
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = green_fill
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.getvalue()


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Return a DataFrame as .xlsx bytes (for st.download_button)."""
    if df is None or df.empty:
        buf = io.BytesIO()
        pd.DataFrame().to_excel(buf, index=False, engine="openpyxl")
        buf.seek(0)
        return buf.getvalue()
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf.getvalue()


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Return a DataFrame as CSV bytes (for st.download_button)."""
    if df is None or df.empty:
        return b""
    return df.to_csv(index=False).encode("utf-8-sig")


def to_excel_bytes_cross_duplicates(cross_df: pd.DataFrame, gtin_outer_col: str = None, gtin_inner_col: str = None) -> bytes:
    """Export Cross Duplicates DataFrame with green fill for all rows (skipped for large sheets)."""
    if cross_df is None or cross_df.empty:
        buf = io.BytesIO()
        pd.DataFrame().to_excel(buf, index=False, engine="openpyxl")
        buf.seek(0)
        return buf.getvalue()
    
    # Prepare display columns
    display_cols = ["Legal Entity"]
    if gtin_outer_col and gtin_outer_col in cross_df.columns:
        display_cols.append(gtin_outer_col)
    if gtin_inner_col and gtin_inner_col in cross_df.columns:
        display_cols.append(gtin_inner_col)
    if "SUPC" in cross_df.columns:
        display_cols.append("SUPC")
    if "Local Product Description" in cross_df.columns:
        display_cols.append("Local Product Description")
    
    # Add normalized columns if they exist
    if "gtin_outer_normalized" in cross_df.columns:
        display_cols.append("gtin_outer_normalized")
    if "gtin_inner_normalized" in cross_df.columns:
        display_cols.append("gtin_inner_normalized")
    
    # Filter to available columns
    available_cols = [col for col in display_cols if col in cross_df.columns]
    export_df = cross_df[available_cols].copy()
    
    # Sort by GTIN for better grouping - use the cross GTIN (appears in both Outer and Inner)
    sort_cols = []
    if "gtin_outer_normalized" in export_df.columns:
        sort_cols.append("gtin_outer_normalized")
    if "Legal Entity" in export_df.columns:
        sort_cols.append("Legal Entity")
    if sort_cols:
        export_df = export_df.sort_values(by=sort_cols).reset_index(drop=True)
    
    buf = io.BytesIO()
    export_df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    
    # Apply green fill if not too large
    if len(export_df) > _INNER_OUTER_STYLE_ROW_LIMIT:
        return buf.getvalue()
    
    wb = load_workbook(buf)
    ws = wb.active
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Apply green fill to all rows (all are Cross Duplicates)
    for row_idx in range(2, len(export_df) + 2):  # Start at row 2 (row 1 is header)
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = green_fill
    
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.getvalue()
